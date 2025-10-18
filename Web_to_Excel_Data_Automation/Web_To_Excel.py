import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# 1. Fetch data from a webpage and convert the annual value to a monthly average
def fetch_annual_value():
    url = "ENTER_THE_WEBPAGE_URL_TO_SCRAPE"  # Replace with the actual data source URL
    headers = {
        "User-Agent": "Mozilla/5.0"
    }
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # Get the first table on the page (assuming it contains the target data)
    table = soup.find("table")
    if table is None:
        raise Exception("Table not found on the page.")

    # Loop through rows to find the first valid annual data
    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) >= 2:
            date_text = tds[0].text.strip()
            annual_value_text = tds[1].text.strip()
            try:
                # Convert comma to dot, cast to float, divide by 12, round to 2 decimals
                value = float(annual_value_text.replace(",", "."))
                monthly_average = round(value / 12, 2)
                return monthly_average
            except ValueError:
                raise Exception(f"Could not convert '{annual_value_text}' to float.")

    raise Exception("Annual value not found in the table.")

# 2. Write the value to a specific cell in an Excel file
def write_to_excel(value):
    file_path = r"C:\Users\Osman ULUHAN\Desktop\Borsa\Algoritmik Trading\250930 Spot & Viop Canlı Algo Takip Tablosu.xlsx"
    sheet_name = "Analiz"

    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise Exception(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    ws["C25"] = value
    wb.save(file_path)